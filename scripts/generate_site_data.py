from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = Path("/Users/daqianlinn/Documents/个人/威尔多芬指标/自如系统净值走势(2026-05-29).xlsx")
OUTPUT_PATH = ROOT / "public" / "data" / "site-data.json"


def to_date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def to_float(value):
    if value is None or value == "-":
        return None
    if isinstance(value, str):
        value = value.replace(",", "").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def pct(value):
    return None if value is None else round(value * 100, 2)


def read_strategy_sheet(ws):
    trades = []
    nav = [{"date": "start", "value": 1.0}]
    for row in ws.iter_rows(min_row=3, values_only=True):
        buy, sell, ret, net = row[:4]
        if not buy or not sell:
            continue
        ret_value = to_float(ret)
        net_value = to_float(net)
        if ret_value is None or net_value is None:
            continue
        item = {
            "buyDate": to_date_text(buy),
            "sellDate": to_date_text(sell),
            "return": ret_value,
            "returnPct": pct(ret_value),
            "net": round(net_value, 4),
            "result": "盈利" if ret_value > 0 else "亏损",
        }
        trades.append(item)
        nav.append({"date": item["sellDate"], "value": item["net"]})
    return trades, nav


def read_index_sheet(ws, buy_label, sell_label):
    series = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        buy, sell, buy_price, sell_price, ret, net, days, gap = row[:8]
        ret_value = to_float(ret)
        net_value = to_float(net)
        if not buy or not sell or ret_value is None or net_value is None:
            continue
        series.append(
            {
                "buyDate": to_date_text(buy),
                "sellDate": to_date_text(sell),
                buy_label: to_float(buy_price),
                sell_label: to_float(sell_price),
                "return": ret_value,
                "returnPct": pct(ret_value),
                "net": round(net_value, 4),
                "holdingDays": int(days) if isinstance(days, (int, float)) else None,
                "waitingDays": int(gap) if isinstance(gap, (int, float)) else None,
            }
        )
    return series


def max_drawdown(nav):
    peak = nav[0]["value"]
    worst = 0
    for point in nav:
        value = point["value"]
        if value > peak:
            peak = value
        if peak:
            worst = min(worst, value / peak - 1)
    return worst


def summarize(trades, nav):
    wins = [x for x in trades if x["return"] > 0]
    losses = [x for x in trades if x["return"] < 0]
    best = max(trades, key=lambda x: x["return"])
    worst = min(trades, key=lambda x: x["return"])
    avg_return = mean([x["return"] for x in trades])
    avg_holding_days = mean(
        [
            (datetime.fromisoformat(x["sellDate"]) - datetime.fromisoformat(x["buyDate"])).days
            for x in trades
        ]
    )
    final_net = nav[-1]["value"]
    return {
        "tradeCount": len(trades),
        "wins": len(wins),
        "losses": len(losses),
        "winRate": round(len(wins) / len(trades), 4),
        "winRatePct": pct(len(wins) / len(trades)),
        "finalNet": round(final_net, 4),
        "cumulativeReturn": round(final_net - 1, 4),
        "cumulativeReturnPct": pct(final_net - 1),
        "averageReturn": round(avg_return, 4),
        "averageReturnPct": pct(avg_return),
        "maxDrawdown": round(max_drawdown(nav), 4),
        "maxDrawdownPct": pct(max_drawdown(nav)),
        "avgHoldingDays": round(avg_holding_days, 0),
        "bestTrade": best,
        "worstTrade": worst,
        "latestClosedTrade": trades[-1],
    }


def build():
    wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    weiler_trades, weiler_nav = read_strategy_sheet(wb["威尔指标"])
    duofen_trades, duofen_nav = read_strategy_sheet(wb["多芬指标"])
    hs300 = read_index_sheet(wb["沪深300"], "buyHs300", "sellHs300")
    cyb = read_index_sheet(wb["创业板指"], "buyChiNext", "sellChiNext")

    data = {
        "meta": {
            "title": "威尔多芬指标",
            "updatedAt": "2026-05-29",
            "sourceWorkbook": WORKBOOK_PATH.name,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "disclaimer": "本页面仅用于指标展示与学习交流，不构成投资建议。",
        },
        "latest": {
            "marketDate": "2026-05-29",
            "weiler": {
                "name": "威尔指标",
                "style": "大盘股指标",
                "benchmark": "沪深300 000300",
                "status": "多头",
                "enteredAt": "2026-04-16",
                "suggestion": "买入或继续持有偏价值、偏大盘风格的基金。",
            },
            "duofen": {
                "name": "多芬指标",
                "style": "小盘股指标",
                "benchmark": "创业板指 399006",
                "status": "多头",
                "enteredAt": "2026-05-28",
                "suggestion": "买入或持有偏成长、偏小盘风格的基金。",
            },
        },
        "strategies": {
            "weiler": {
                "name": "威尔指标",
                "shortName": "威尔",
                "style": "大盘价值",
                "color": "#f97316",
                "benchmark": "沪深300",
                "description": "威尔主要观察大盘股、价值股的机会，适合用来判断大盘价值风格是否值得进攻。",
                "trades": weiler_trades,
                "nav": weiler_nav,
                "benchmarkSeries": [{"date": x["sellDate"], "value": x["net"]} for x in hs300],
                "metrics": summarize(weiler_trades, weiler_nav),
            },
            "duofen": {
                "name": "多芬指标",
                "shortName": "多芬",
                "style": "小盘成长",
                "color": "#3b82f6",
                "benchmark": "创业板指",
                "description": "多芬主要观察小盘股、成长股的机会，适合用来判断小盘成长风格是否值得进攻。",
                "trades": duofen_trades,
                "nav": duofen_nav,
                "benchmarkSeries": [{"date": x["sellDate"], "value": x["net"]} for x in cyb],
                "metrics": summarize(duofen_trades, duofen_nav),
            },
        },
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    build()
